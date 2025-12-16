import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import time
import threading
from datetime import datetime
from TCPClient import TCPClient
from SimulatorMessageHandler import CtrlMessageHandler, StatusMessageHandler
from WaveformWindow import WaveformWindow
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


class HardwareSimulator:
    def __init__(self, root):
        self.root = root
        self.root.title("硬件仿真系统 v1.0.0")
        # 获取屏幕尺寸并设置窗口大小
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        print(f"{screen_width}x{screen_height}")

        # 设置窗口为屏幕大小，并定位到左上角
        self.root.geometry(f"{screen_width}x{screen_height}")

        # 设置整体背景色为灰色
        self.root.configure(bg='#d9d9d9')
        self.query_index = 1

        # 状态变量
        self.model_file = None
        self.is_running = False
        self.start_time = None
        self.timer_thread = None
        self.stop_timer = False
        self.is_connected = False

        # 心跳控制
        self.use_heartbeat = True  # 是否使用心跳机制
        self.last_heartbeat_time = None  # 最后收到心跳的时间
        self.heartbeat_timer = None  # 心跳定时器
        self.heartbeat_interval = 5  # 心跳发送间隔（秒）
        self.heartbeat_timeout = 20  # 心跳超时时间（秒）
        self.status_check_timer = None  # 状态检查定时器

        # 消息处理器
        self.ctrl_handler = None
        self.status_handler = None

        # 查询定时器控制变量
        self.query_timer = None
        self.is_querying = False
        self.query_interval = 1.0  # 查询间隔1秒

        # 波形窗口管理
        self.waveform_windows = {}  # 存储打开的波形窗口

        # 创建示例JSON文件（如果不存在）
        self.create_sample_json_files()

        # 加载参数和变量数据
        self.input_params = self.load_json_file("input_params.json")
        self.watch_variables = self.load_json_file("watch_variables.json")

        # 保存原始参数值（用于比较修改）
        self.original_input_params = [param.copy() for param in self.input_params]

        # 存储表格行引用
        self.param_rows = []
        self.watch_rows = []

        self.create_widgets()
        self.add_log("系统启动成功...")

        # 数据记录相关
        self.data_record_file = "仿真数据记录.xlsx"
        self.param_record_count = 0
        self.var_record_count = 0
        self.save_data_excel = True  # 数据保存开关，True时保存到Excel，False时不保存
        self.init_data_record_file()  # 初始化Excel文件

        # 初始化修改状态显示
        if hasattr(self, '_update_modified_status'):
            self._update_modified_status()

    def create_sample_json_files(self):
        """创建示例JSON文件"""
        try:
            # 检查文件是否已存在
            with open("input_params.json", "r", encoding="utf-8") as f:
                json.load(f)  # 尝试读取
        except (FileNotFoundError, json.JSONDecodeError):
            # 创建输入参数示例文件
            input_params = [
                {"param": "param1", "type": "float", "val": "1000"},
                {"param": "param2", "type": "float", "val": "10.5"},
                {"param": "param3", "type": "float", "val": "100"},
                {"param": "param4", "type": "float", "val": "50"},
                {"param": "param5", "type": "float", "val": "3.14"},
                {"param": "param6", "type": "float", "val": "0.1"}
            ]

            with open("input_params.json", "w", encoding="utf-8") as f:
                json.dump(input_params, f, ensure_ascii=False, indent=2)
            print("创建了 input_params.json 文件")

        try:
            with open("watch_variables.json", "r", encoding="utf-8") as f:
                json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            # 创建监视变量示例文件
            watch_variables = [
                {"variable": "var1", "type": "float", "val": "25.5"},
                {"variable": "var2", "type": "float", "val": "101.3"},
                {"variable": "var3", "type": "float", "val": "1500"}
            ]

            with open("watch_variables.json", "w", encoding="utf-8") as f:
                json.dump(watch_variables, f, ensure_ascii=False, indent=2)
            print("创建了 watch_variables.json 文件")

    def load_json_file(self, filename):
        """加载JSON文件，如果文件不存在则返回空列表"""
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(f"成功加载 {filename}: {len(data)} 条数据")
                return data
        except FileNotFoundError:
            print(f"警告: 文件 {filename} 不存在")
            return []
        except json.JSONDecodeError as e:
            print(f"错误: 文件 {filename} JSON格式错误: {e}")
            return []

    def init_data_record_file(self):
        """初始化数据记录Excel文件"""
        try:
            # 检查文件是否已存在
            if os.path.exists(self.data_record_file):
                self.add_log(f"数据记录文件已存在: {self.data_record_file}")
                return

            # 创建新的Excel文件
            with pd.ExcelWriter(self.data_record_file, engine='openpyxl') as writer:
                # 创建输入参数表格
                param_headers = ["编号"]
                for i, param in enumerate(self.input_params, 1):
                    param_name = param.get("param", f"param{i}")
                    param_headers.append(param_name)
                param_headers.append("时间")

                # 创建空的DataFrame
                param_df = pd.DataFrame(columns=param_headers)
                param_df.to_excel(writer, sheet_name='输入参数', index=False)

                # 创建观察变量表格
                var_headers = ["编号"]
                for i, var in enumerate(self.watch_variables, 1):
                    var_name = var.get("variable", f"var{i}")
                    var_headers.append(var_name)
                var_headers.append("时间")

                var_df = pd.DataFrame(columns=var_headers)
                var_df.to_excel(writer, sheet_name='观察变量', index=False)

                # 设置列格式
                workbook = writer.book
                worksheet = workbook['观察变量']

                # 设置编号列为整数格式
                worksheet.column_dimensions['A'].width = 10

                # 设置时间列格式
                time_col = get_column_letter(len(var_headers))
                worksheet.column_dimensions[time_col].width = 20

                # 设置变量值列格式
                for i, var in enumerate(self.watch_variables, 2):  # 从第2列开始
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = 15

                    # 根据变量类型设置格式
                    var_type = var.get("type", "float")
                    if var_type == "int":
                        # 整数格式
                        for row in range(2, 1002):  # 预设置格式
                            worksheet.cell(row=row, column=i).number_format = '0'
                    else:
                        # 浮点数格式，统一3位小数
                        for row in range(2, 1002):  # 预设置格式
                            worksheet.cell(row=row, column=i).number_format = '0.000'

            self.add_log(f"已创建数据记录文件: {self.data_record_file}")

        except Exception as e:
            self.add_log(f"初始化数据记录文件失败: {e}")

    def record_parameters(self, params_data, timestamp=None):
        """记录参数数据到Excel文件"""
        try:
            # 检查数据保存开关
            if not self.save_data_excel:
                return

            if timestamp is None:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 构建数据行
            self.param_record_count += 1
            row_data = [self.param_record_count]

            # 按参数顺序添加值
            for param in self.input_params:
                param_name = param.get("param", "")
                param_value = params_data.get(param_name, "")
                row_data.append(param_value)

            row_data.append(timestamp)

            # 读取现有数据
            try:
                existing_df = pd.read_excel(self.data_record_file, sheet_name='输入参数')
            except:
                existing_df = pd.DataFrame()

            # 添加新行
            headers = ["编号"]
            for param in self.input_params:
                headers.append(param.get("param", ""))
            headers.append("时间")

            new_row = pd.DataFrame([row_data], columns=headers)
            updated_df = pd.concat([existing_df, new_row], ignore_index=True)

            # 保存到Excel
            with pd.ExcelWriter(self.data_record_file, engine='openpyxl', mode='a',
                                if_sheet_exists='replace') as writer:
                updated_df.to_excel(writer, sheet_name='输入参数', index=False)

            self.add_log(f"已记录参数数据，编号: {self.param_record_count}")

        except Exception as e:
            self.add_log(f"记录参数数据失败: {e}")

    def record_variables(self, vars_data, timestamp=None):
        """记录变量数据到Excel文件"""
        try:
            # 检查数据保存开关
            if not self.save_data_excel:
                return

            if timestamp is None:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 使用openpyxl直接写入，保持格式
            from openpyxl import load_workbook

            # 加载工作簿
            workbook = load_workbook(self.data_record_file)
            worksheet = workbook['观察变量']

            # 找到最后一行
            last_row = worksheet.max_row + 1
            self.var_record_count += 1

            # 写入编号
            worksheet.cell(row=last_row, column=1, value=self.var_record_count)

            # 按变量顺序写入值
            for i, var in enumerate(self.watch_variables, 2):  # 从第2列开始
                var_name = var.get("variable", "")
                var_value = vars_data.get(var_name, "")

                if var_value is not None and var_value != "":
                    try:
                        # 尝试转换为数值
                        num_value = float(var_value)
                        var_type = var.get("type", "float")

                        if var_type == "int":
                            worksheet.cell(row=last_row, column=i, value=int(num_value))
                        else:
                            # 浮点数保留3位小数
                            worksheet.cell(row=last_row, column=i, value=round(num_value, 3))
                    except (ValueError, TypeError):
                        # 如果转换失败，使用原始字符串
                        worksheet.cell(row=last_row, column=i, value=str(var_value))
                else:
                    worksheet.cell(row=last_row, column=i, value="")

            # 写入时间
            time_col = len(self.watch_variables) + 2
            worksheet.cell(row=last_row, column=time_col, value=timestamp)

            # 保存工作簿
            workbook.save(self.data_record_file)

            self.add_log(f"已记录变量数据，编号: {self.var_record_count}")

        except Exception as e:
            self.add_log(f"记录变量数据失败: {e}")

    def create_widgets(self):
        # 主标题
        title_label = tk.Label(self.root, text="硬件仿真系统 v1.0.0",
                               font=("Arial", 16, "bold"), bg='#d9d9d9')
        title_label.pack(pady=10)

        # 创建主容器 - 背景色设置为灰色
        main_container = tk.Frame(self.root, bg='#d9d9d9')
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        # 1. 基础设置区域
        basic_settings_frame = tk.LabelFrame(main_container, text="基础设置", font=("Arial", 11, "bold"),
                                             bg='#d9d9d9', fg='#333333', bd=2, relief=tk.GROOVE)
        basic_settings_frame.pack(fill=tk.X, pady=(0, 15))

        # 目标机区域（第一行）
        target_frame = tk.Frame(basic_settings_frame, bg='#d9d9d9')
        target_frame.pack(fill=tk.X, pady=(10, 5), anchor='w', padx=10)

        tk.Label(target_frame, text="目标机:", font=("Arial", 10), bg='#d9d9d9').pack(side=tk.LEFT, padx=(0, 5))

        self.target_entry = tk.Entry(target_frame, width=20, font=("Arial", 10), bg='white',
                                     highlightthickness=0, bd=3, relief='flat',
                                     highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.target_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.target_entry.insert(0, "192.168.3.173")

        self.connect_button = tk.Button(target_frame, text="连接", width=8, font=("Arial", 10),
                                        command=self.toggle_connection, bg='#d9d9d9',
                                        highlightthickness=0, bd=0, relief='flat',
                                        highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.connect_button.pack(side=tk.LEFT)

        # 添加连接状态显示
        self.status_frame = tk.Frame(target_frame, bg='#d9d9d9')
        self.status_frame.pack(side=tk.LEFT)

        # 状态圆圈
        self.status_canvas = tk.Canvas(self.status_frame, width=20, height=20, bg='#d9d9d9',
                                       highlightthickness=0)
        self.status_canvas.pack(side=tk.LEFT, padx=(0, 5))
        self.status_circle = self.status_canvas.create_oval(5, 5, 15, 15, fill='red', outline='')

        # 状态文字
        self.status_label = tk.Label(self.status_frame, text="目标机离线", font=("Arial", 10),
                                     bg='#d9d9d9', fg='red')
        self.status_label.pack(side=tk.LEFT)

        # 模型操作区域（第二行）
        model_ops_frame = tk.Frame(basic_settings_frame, bg='#d9d9d9')
        model_ops_frame.pack(fill=tk.X, pady=(5, 10), anchor='w', padx=10)

        # 选择模型按钮和文件名显示
        self.select_model_btn = tk.Button(model_ops_frame, text="选择模型", width=10, font=("Arial", 10),
                                          command=self.select_model, bg='#d9d9d9',
                                          highlightthickness=0, bd=0, relief='flat',
                                          highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.select_model_btn.pack(side=tk.LEFT, padx=(0, 10))

        # 模型文件名显示
        self.model_file_label = tk.Label(model_ops_frame, text="未选择文件", font=("Arial", 10),
                                         bg='#d9d9d9', width=15, anchor='w')
        self.model_file_label.pack(side=tk.LEFT, padx=(0, 20))

        # 模型下载按钮和状态显示
        self.download_model_btn = tk.Button(model_ops_frame, text="模型下载", width=10, font=("Arial", 10),
                                            command=self.download_model, bg='#d9d9d9',
                                            highlightthickness=0, bd=0, relief='flat',
                                            highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.download_model_btn.pack(side=tk.LEFT, padx=(0, 10))

        # 模型下载状态显示
        self.download_status_label = tk.Label(model_ops_frame, text="未下载", font=("Arial", 10),
                                              bg='#d9d9d9', width=12, anchor='w')
        self.download_status_label.pack(side=tk.LEFT, padx=(0, 20))

        # 模型运行按钮和时间显示
        self.run_button = tk.Button(model_ops_frame, text="模型运行", width=10, font=("Arial", 10),
                                    command=self.toggle_model_run, bg='#d9d9d9',
                                    highlightthickness=0, bd=0, relief='flat',
                                    highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.run_button.pack(side=tk.LEFT, padx=(0, 10))

        # 运行时间显示
        self.time_label = tk.Label(model_ops_frame, text="00:00:00", font=("Arial", 10),
                                   bg='#d9d9d9', width=10, anchor='center')
        self.time_label.pack(side=tk.LEFT)

        # 2. 参数变量区域
        params_vars_frame = tk.LabelFrame(main_container, text="参数变量", font=("Arial", 11, "bold"),
                                          bg='#d9d9d9', fg='#333333', bd=2, relief=tk.GROOVE)
        params_vars_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # 表格标题行
        table_titles_frame = tk.Frame(params_vars_frame, bg='#d9d9d9')
        table_titles_frame.pack(fill=tk.X, pady=(10, 5), padx=10)

        # 左侧：模型参数输入标题 + 参数下发按钮
        left_title_frame = tk.Frame(table_titles_frame, bg='#d9d9d9')
        left_title_frame.pack(side=tk.LEFT, anchor='w')

        tk.Label(left_title_frame, text="模型参数输入", font=("Arial", 10, "bold"), bg='#d9d9d9').pack(side=tk.LEFT,
                                                                                                       padx=(0, 10))

        self.param_send_btn = tk.Button(left_title_frame, text="参数下发", width=10, font=("Arial", 10),
                                        command=self.send_parameters, bg='#d9d9d9',
                                        highlightthickness=0, bd=0, relief='flat',
                                        highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.param_send_btn.pack(side=tk.LEFT, padx=(20, 20))

        # 添加"参数复位"按钮
        self.param_reset_btn = tk.Button(left_title_frame, text="参数复位", width=10, font=("Arial", 10),
                                         command=self.reset_parameters, bg='#d9d9d9',
                                         highlightthickness=0, bd=0, relief='flat',
                                         highlightbackground='#d9d9d9', highlightcolor='#d9d9d9')
        self.param_reset_btn.pack(side=tk.LEFT)

        # 修改状态标签
        self.modified_label = tk.Label(left_title_frame, text="", font=("Arial", 9),
                                       bg='#d9d9d9', fg='red')
        self.modified_label.pack(side=tk.LEFT, padx=(30, 10))

        # 右侧：变量监视标题
        right_title_frame = tk.Frame(table_titles_frame, bg='#d9d9d9')
        right_title_frame.pack(side=tk.RIGHT, anchor='w')

        tk.Label(right_title_frame, text="变量监视", font=("Arial", 10, "bold"), bg='#d9d9d9').pack(side=tk.LEFT)

        # 表格容器
        tables_container = tk.Frame(params_vars_frame, bg='#d9d9d9')
        tables_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        # 左侧参数表格
        params_frame = tk.Frame(tables_container, bg='#d9d9d9')
        params_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        self.setup_params_table(params_frame)

        # 右侧监视表格
        watch_frame = tk.Frame(tables_container, bg='#d9d9d9')
        watch_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        self.setup_watch_table(watch_frame)

        # 3. 底部：系统日志区域
        system_log_frame = tk.LabelFrame(main_container, text="系统日志", font=("Arial", 11, "bold"),
                                         bg='#d9d9d9', fg='#333333', bd=2, relief=tk.GROOVE)
        system_log_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # 日志文本框容器
        log_container = tk.Frame(system_log_frame, bg='#d9d9d9')
        log_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 日志文本框
        log_frame = tk.Frame(log_container, bg='#d9d9d9')
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, wrap=tk.WORD, font=("Arial", 10), height=8, bg='white')
        log_scrollbar = tk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview, width=3)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)

        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_params_table(self, parent):
        """设置参数输入表格 - 使用Frame和Label创建真实表格"""
        # 创建带边框的表格框架
        table_frame = tk.Frame(parent, bg='black', bd=1, relief='solid')
        table_frame.pack(fill=tk.BOTH, expand=True)

        # 创建表头
        header_frame = tk.Frame(table_frame, bg='lightgray')
        header_frame.pack(fill=tk.X)

        # 表头标签 - 使用固定宽度
        headers = ["索引", "输入参数", "参数数值"]
        widths = [10, 48, 48]  # 字符宽度

        header_labels = []
        for i, (header, width) in enumerate(zip(headers, widths)):
            label = tk.Label(header_frame, text=header, font=('Arial', 10, 'bold'),
                             bg='lightgray', width=width, relief='solid', bd=1,
                             anchor=tk.CENTER)
            if i == len(headers) - 1:  # 最后一列填充剩余空间
                label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            else:
                label.pack(side=tk.LEFT, fill=tk.BOTH)
            header_labels.append(label)

        # 存储表头宽度信息
        self.param_header_widths = widths

        # 创建表格内容框架（带滚动条）
        content_frame = tk.Frame(table_frame, bg='white')
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 创建Canvas和Scrollbar用于滚动
        canvas = tk.Canvas(content_frame, bg='white', highlightthickness=0)
        scrollbar = tk.Scrollbar(content_frame, orient=tk.VERTICAL, command=canvas.yview, width=3)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 存储表格行引用
        self.param_rows = []

        # 配置画布滚动
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _on_mousewheel)
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.params_content_frame = scrollable_frame
        self.params_canvas = canvas

        # 填充数据
        self.update_params_table()

    def setup_watch_table(self, parent):
        """设置变量监视表格 - 使用Frame和Label创建真实表格"""
        # 创建带边框的表格框架
        table_frame = tk.Frame(parent, bg='black', bd=1, relief='solid')
        table_frame.pack(fill=tk.BOTH, expand=True)

        # 创建表头
        header_frame = tk.Frame(table_frame, bg='lightgray')
        header_frame.pack(fill=tk.X)

        # 表头标签 - 使用固定宽度
        headers = ["索引", "变量名称", "参数数值", "波形"]
        widths = [10, 50, 50, 7]  # 字符宽度

        header_labels = []
        for i, (header, width) in enumerate(zip(headers, widths)):
            label = tk.Label(header_frame, text=header, font=('Arial', 10, 'bold'),
                             bg='lightgray', width=width, relief='solid', bd=1,
                             anchor=tk.CENTER)
            if i == len(headers) - 1:  # 最后一列填充剩余空间
                label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            else:
                label.pack(side=tk.LEFT, fill=tk.BOTH)
            header_labels.append(label)

        # 存储表头宽度信息
        self.watch_header_widths = widths

        # 创建表格内容框架（带滚动条）
        content_frame = tk.Frame(table_frame, bg='white')
        content_frame.pack(fill=tk.BOTH, expand=True)

        # 创建Canvas和Scrollbar用于滚动
        canvas = tk.Canvas(content_frame, bg='white', highlightthickness=0)
        scrollbar = tk.Scrollbar(content_frame, orient=tk.VERTICAL, command=canvas.yview, width=3)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 存储表格行引用
        self.watch_rows = []

        # 配置画布滚动
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _on_mousewheel)
        scrollable_frame.bind("<MouseWheel>", _on_mousewheel)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.watch_content_frame = scrollable_frame
        self.watch_canvas = canvas

        # 填充数据
        self.update_watch_table()

    def update_params_table(self):
        """更新参数表格"""
        # 清空现有行
        for widget in self.params_content_frame.winfo_children():
            widget.destroy()

        self.param_rows = []

        # 添加数据行
        for idx, param in enumerate(self.input_params, 1):
            param_name = param.get("param", f"参数{idx}")
            param_value = param.get("val", "")

            # 创建一行
            row_frame = tk.Frame(self.params_content_frame, bg='white')
            row_frame.pack(fill=tk.X)

            # 索引列
            index_label = tk.Label(row_frame, text=str(idx), font=('Arial', 10),
                                   bg='white', width=self.param_header_widths[0],
                                   height=2, relief='solid', bd=1, anchor=tk.CENTER)
            index_label.pack(side=tk.LEFT, fill=tk.BOTH)

            # 参数名列
            name_label = tk.Label(row_frame, text=param_name, font=('Arial', 10),
                                  bg='white', width=self.param_header_widths[1],
                                  height=2, relief='solid', bd=1, anchor=tk.CENTER)
            name_label.pack(side=tk.LEFT, fill=tk.BOTH)

            # 参数值列
            value_label = tk.Label(row_frame, text=param_value, font=('Arial', 10),
                                   bg='white', width=self.param_header_widths[2],
                                   height=2, relief='solid', bd=1, anchor=tk.CENTER, cursor="hand2")
            value_label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # 绑定双击编辑事件
            value_label.bind("<Double-1>",
                             lambda e, idx=idx - 1, label=value_label:
                             self.edit_param_value(idx, label))

            # 存储行引用
            self.param_rows.append({
                'frame': row_frame,
                'index': index_label,
                'name': name_label,
                'value': value_label
            })

        # 更新滚动区域
        self.params_content_frame.update_idletasks()
        self.params_canvas.configure(scrollregion=self.params_canvas.bbox("all"))

    def update_watch_table(self):
        """更新监视变量表格"""
        # 清空现有行
        for widget in self.watch_content_frame.winfo_children():
            widget.destroy()

        self.watch_rows = []

        # 添加数据行
        for idx, var in enumerate(self.watch_variables, 1):
            var_name = var.get("variable", f"变量{idx}")
            var_value = var.get("val", "")

            # 创建一行
            row_frame = tk.Frame(self.watch_content_frame, bg='white')
            row_frame.pack(fill=tk.X)

            # 索引列
            index_label = tk.Label(row_frame, text=str(idx), font=('Arial', 10),
                                   bg='white', width=self.watch_header_widths[0],
                                   height=2, relief='solid', bd=1, anchor=tk.CENTER)
            index_label.pack(side=tk.LEFT, fill=tk.BOTH)

            # 变量名列
            name_label = tk.Label(row_frame, text=var_name, font=('Arial', 10),
                                  bg='white', width=self.watch_header_widths[1],
                                  height=2, relief='solid', bd=1, anchor=tk.CENTER)
            name_label.pack(side=tk.LEFT, fill=tk.BOTH)

            # 变量值列
            value_label = tk.Label(row_frame, text=var_value, font=('Arial', 10),
                                   bg='white', width=self.watch_header_widths[2],
                                   height=2, relief='solid', bd=1, anchor=tk.CENTER)
            value_label.pack(side=tk.LEFT, fill=tk.BOTH)

            # 波形按钮列
            wave_button = tk.Label(row_frame, text="＿/￣", font=('Arial', 8),
                                   bg='lightblue', width=self.watch_header_widths[3],
                                   height=2, relief='raised', bd=1, anchor=tk.CENTER, cursor="hand2")
            wave_button.pack(side=tk.LEFT, fill=tk.BOTH)

            # 绑定波形按钮点击事件
            wave_button.bind("<Button-1>",
                             lambda e, name=var_name: self.on_waveform_click_new(name))

            # 存储行引用
            self.watch_rows.append({
                'frame': row_frame,
                'index': index_label,
                'name': name_label,
                'value': value_label,
                'wave': wave_button
            })

        # 更新滚动区域
        self.watch_content_frame.update_idletasks()
        self.watch_canvas.configure(scrollregion=self.watch_canvas.bbox("all"))

    def edit_param_value(self, idx, label):
        """编辑参数值"""
        current_value = label.cget("text")

        # 记录原始值用于比较
        original_value = self.input_params[idx].get("val", "")

        # 创建编辑窗口
        edit_win = tk.Toplevel(self.root)
        edit_win.title("编辑参数值")
        edit_win.geometry("350x150")
        edit_win.transient(self.root)
        edit_win.grab_set()

        tk.Label(edit_win, text=f"参数: {self.input_params[idx].get('param', f'参数{idx + 1}')}",
                 font=("Arial", 10, "bold")).pack(pady=(10, 5))

        # tk.Label(edit_win, text="输入新值:").pack()

        entry = tk.Entry(edit_win, width=30)
        entry.insert(0, current_value)
        entry.pack(pady=5)
        entry.focus_set()
        entry.select_range(0, tk.END)

        # 显示原始值
        # tk.Label(edit_win, text=f"原始值: {original_value}",
        #          font=("Arial", 9), fg="gray").pack(pady=5)

        def save_edit():
            new_value = entry.get()
            label.config(text=new_value)

            # 更新内存中的数据
            if 0 <= idx < len(self.input_params):
                # 记录旧值
                old_value = self.input_params[idx].get("val", "")
                # 更新为新值
                self.input_params[idx]["val"] = new_value

                # 修改日志：记录从什么值修改为什么值
                param_name = self.input_params[idx].get("param", f"参数{idx + 1}")
                if old_value != new_value:
                    self.add_log(f"参数 '{param_name}' 值由 '{old_value}' 修改为 '{new_value}'")
                else:
                    self.add_log(f"参数 '{param_name}' 值未改变: {new_value}")

            # 更新修改状态显示
            self._update_modified_status()

            edit_win.destroy()

        button_frame = tk.Frame(edit_win)
        button_frame.pack(pady=10)

        tk.Button(button_frame, text="确定", width=8, command=save_edit).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="取消", width=8, command=edit_win.destroy).pack(side=tk.LEFT, padx=5)

        entry.bind("<Return>", lambda e: save_edit())

    def on_waveform_click_new(self, variable_name):
        """波形按钮点击事件"""
        self.add_log(f"点击了变量 '{variable_name}' 的波形按钮")
        # messagebox.showinfo("波形显示", f"显示变量 {variable_name} 的波形")
        # 检查是否已经打开了该变量的波形窗口
        if variable_name in self.waveform_windows:
            window = self.waveform_windows[variable_name]
            if window.is_open():
                # 窗口已存在，将其提到前台
                window.window.lift()
                window.window.focus_force()
                return

        # 创建新的波形窗口
        try:
            wave_window = WaveformWindow(self.root, variable_name, max_points=200)
            self.waveform_windows[variable_name] = wave_window

            # 绑定窗口关闭事件
            wave_window.window.protocol("WM_DELETE_WINDOW",
                                        lambda v=variable_name: self._on_waveform_window_close(v))

        except Exception as e:
            self.add_log(f"创建波形窗口失败: {e}")
            messagebox.showerror("错误", f"无法创建波形窗口: {e}")

    def _on_waveform_window_close(self, variable_name):
        """波形窗口关闭时的处理"""
        if variable_name in self.waveform_windows:
            try:
                # 销毁窗口
                window = self.waveform_windows[variable_name]
                if window.is_open():
                    window.destroy()
            except:
                pass
            finally:
                # 从字典中移除
                del self.waveform_windows[variable_name]

    def _update_waveform_data(self, variable_name, value):
        """更新波形窗口数据"""
        try:
            if variable_name in self.waveform_windows:
                window = self.waveform_windows[variable_name]
                if window.is_open():
                    window.add_data_point(value)
        except Exception as e:
            # 静默处理错误，避免影响主程序
            pass

    def toggle_connection(self):
        """切换连接状态"""
        target = self.target_entry.get()
        if not self.is_connected:
            # 连接操作
            self.connect_button.config(text="连接中...", state="disabled")

            # 更新连接状态显示
            self._update_connection_status_display(False, force_disconnect=True)

            def connect_thread():
                try:
                    # 创建消息处理器，使用新的回调设计
                    self.ctrl_handler = CtrlMessageHandler(
                        target, 9001,
                        message_callback=self.on_system_message
                    )
                    self.status_handler = StatusMessageHandler(
                        target, 9000,
                        variable_callback=self.on_variable_data
                    )

                    # 启动处理器
                    ctrl_success = self.ctrl_handler.start()
                    status_success = self.status_handler.start()

                    # 更新UI
                    self.root.after(0, self._update_connection_status, target, ctrl_success, status_success)

                except Exception as e:
                    self.root.after(0, lambda: self.add_log(f"连接过程中发生错误: {e}"))
                    self.root.after(0, lambda: self.connect_button.config(text="连接", state="normal"))
                    self.root.after(0, lambda: self._update_connection_status_display(False, force_disconnect=True))

            threading.Thread(target=connect_thread, daemon=True).start()
        else:
            # 断开连接操作
            self._disconnect_connections(force_status_update=True)
            self.add_log(f"已断开与目标机 {target} 的连接")

    def on_system_message(self, message_info):
        """
        处理系统控制消息（来自控制链路）
        Args:
            message_info: 包含系统消息信息的字典
        """
        # 在UI线程中处理系统消息
        self.root.after(0, lambda: self._handle_system_message_ui(message_info))

    def on_variable_data(self, variable_info):
        """
        处理变量数据（来自状态链路）
        Args:
            variable_info: 包含变量数据的字典
        """
        # 在UI线程中更新变量显示
        self.root.after(0, lambda: self._handle_variable_data_ui(variable_info))

    def _handle_system_message_ui(self, message_info):
        """在UI线程中处理系统消息"""
        try:
            data = json.loads(message_info)
            cmd = data.get('cmd', 'unknown')
            if cmd == 'Heart_ack':
                self._handle_heartbeat_message(data)

        except Exception as e:
            self.add_log(f"处理系统消息UI错误: {e}")

    def _handle_variable_data_ui(self, variable_info):
        """在UI线程中处理变量数据"""
        try:
            # 应该判断cmd类型
            self._update_variables_from_data(variable_info)

            # 检查是否需要更新波形窗口
            vars_dict = variable_info.get('vars', {})
            if vars_dict:
                # 获取时间戳
                timestamp_ms = variable_info.get('time', None)
                if timestamp_ms:
                    # 将毫秒时间戳转换为日期时间
                    try:
                        timestamp_sec = timestamp_ms / 1000.0
                        dt = datetime.fromtimestamp(timestamp_sec)
                        time_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                    except:
                        time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                else:
                    time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # 记录变量数据到Excel
                try:
                    self.record_variables(vars_dict, time_str)
                except Exception as e:
                    self.add_log(f"记录变量数据失败: {e}")

                for var_name, var_value in vars_dict.items():
                    if var_name in self.waveform_windows:
                        window = self.waveform_windows[var_name]
                        if window.is_open():
                            # 在UI线程中更新波形
                            self.root.after(0, lambda vn=var_name, vv=var_value:
                            self._update_waveform_data(vn, vv))

        except Exception as e:
            self.add_log(f"处理变量数据UI错误: {e}")

    def _parse_query_var_message(self, message):
        try:
            # 尝试解析JSON
            # data = json.loads(message)
            data = message

            # 检查是否为查询变量响应
            if data.get('cmd') == 'QueryVars_ack':
                ack = data.get('ack', '')
                if ack == 'OK':
                    return data
                else:
                    self.add_log(f"查询变量响应失败: {ack}")
            else:
                # 如果不是QueryVars_ack，记录日志但不处理
                self.add_log(f"收到非查询变量响应: {data.get('cmd')}")

        except json.JSONDecodeError as e:
            self.add_log(f"解析JSON失败: {e}, 消息: {message[:100]}...")
        except Exception as e:
            self.add_log(f"解析查询变量消息失败: {e}")

        return None

    def _update_variables_from_data(self, data):
        """从解析的数据更新变量"""
        try:
            vars_dict = data.get('vars', {})
            if not vars_dict:
                return

            updated = False
            for var_name, var_value in vars_dict.items():
                # 在监视变量列表中查找匹配的变量
                for var in self.watch_variables:
                    if var.get('variable') == var_name:
                        # 统一处理数值格式
                        old_value = var.get('val', '')

                        # 确保新值格式统一
                        if var_value is not None:
                            try:
                                # 根据变量类型决定格式
                                var_type = var.get("type", "float")
                                if var_type == "int":
                                    new_value = str(int(float(var_value)))
                                else:
                                    # 浮点数统一保留3位小数
                                    new_value = f"{float(var_value):.3f}"
                            except (ValueError, TypeError):
                                new_value = str(var_value)
                        else:
                            new_value = ""

                        var['val'] = new_value

                        # 记录变化
                        if old_value != new_value:
                            self.add_log(f"变量更新: {var_name} = {new_value}")
                            updated = True
                        break

            # 如果有变量被更新，刷新表格显示
            if updated:
                self.update_watch_table()
                self.add_log(f"已更新{len(vars_dict)}个变量")

        except Exception as e:
            self.add_log(f"更新变量失败: {e}")

    def _handle_heartbeat_message(self, data):
        """处理心跳响应消息"""
        print('_handle_heartbeat_message')
        try:
            # 更新最后收到心跳的时间
            self.last_heartbeat_time = time.time()

            # 获取心跳时间戳
            act_time = data.get('act', '')
            if act_time:
                self.add_log(f"收到心跳响应，服务器时间: {act_time}")
            else:
                self.add_log("收到心跳响应")

            # 更新连接状态显示
            self._update_connection_status_display(True)

        except Exception as e:
            self.add_log(f"处理心跳消息失败: {e}")

    def _handle_connected_message(self):
        """处理连接成功消息"""
        # 更新最后心跳时间
        self.last_heartbeat_time = time.time()

        # 更新连接状态显示
        self._update_connection_status_display(True)

        # 启动心跳机制
        if self.use_heartbeat:
            self._start_heartbeat_mechanism()

    def _handle_error_message(self, message):
        """处理错误消息"""
        self.add_log(f"连接错误: {message}")
        self._update_connection_status_display(False)

    def _update_connection_status(self, target, ctrl_success, status_success):
        """更新连接状态"""
        if ctrl_success and status_success:
            self.is_connected = True
            self.connect_button.config(text="断开", bg="lightcoral", state="normal")
            self.add_log(f"控制链路(9001)连接成功 - 目标机: {target}")
            self.add_log(f"状态链路(9000)连接成功 - 目标机: {target}")
            self.add_log(f"已连接到目标机: {target}")

            # 更新最后心跳时间
            self.last_heartbeat_time = time.time()

            # 更新连接状态显示
            self._update_connection_status_display(True, force_disconnect=True)

            # 启动心跳机制
            if self.use_heartbeat:
                self._start_heartbeat_mechanism()

        else:
            self.is_connected = False
            self.connect_button.config(text="连接", bg="SystemButtonFace", state="normal")

            # 更新连接状态显示
            self._update_connection_status_display(False, force_disconnect=True)

            # 停止心跳机制
            self._stop_heartbeat_mechanism()

            if not ctrl_success and not status_success:
                self.add_log(f"连接失败: 控制链路(9001)和状态链路(9000)都无法连接到目标机 {target}")
            elif not ctrl_success:
                self.add_log(f"连接失败: 控制链路(9001)无法连接到目标机 {target}")
                if self.status_handler:
                    self.status_handler.stop()
            else:
                self.add_log(f"控制链路(9001)连接成功 - 目标机: {target}")
                self.add_log(f"连接失败: 状态链路(9000)无法连接到目标机 {target}")
                if self.ctrl_handler:
                    self.ctrl_handler.stop()

            self.ctrl_handler = None
            self.status_handler = None

    def _disconnect_connections(self, force_status_update=False):
        """断开所有连接

        Args:
            force_status_update: 是否强制更新状态显示
        """
        # 停止查询定时器
        self._stop_var_query_timer()

        # 停止心跳机制
        self._stop_heartbeat_mechanism()

        # 关闭所有波形窗口
        for variable_name in list(self.waveform_windows.keys()):
            self._on_waveform_window_close(variable_name)

        if self.ctrl_handler:
            self.ctrl_handler.stop()
            self.ctrl_handler = None
        if self.status_handler:
            self.status_handler.stop()
            self.status_handler = None

        # 重置连接状态
        self.is_connected = False
        self.last_heartbeat_time = None

        # 更新连接状态显示
        self.root.after(0, lambda: self._update_connection_status_display(False, force_disconnect=force_status_update))

        # 更新按钮状态
        self.root.after(0, lambda: self.connect_button.config(text="连接", bg="SystemButtonFace", state="normal"))

    def _send_query_var_message(self):
        """发送查询变量消息"""
        print('_send_query_var_message');
        if self.status_handler and self.status_handler.is_connected():
            # 构建查询消息
            query_data = {
                "cmd": "QueryVars",
                "count": len(self.watch_variables)
            }
            json_str = json.dumps(query_data)

            if self.status_handler.send_message(json_str):
                self.add_log(f"发送变量查询: {json_str}")
            else:
                self.add_log("变量查询发送失败")
            self.query_index += 1

    def _start_var_query_timer(self):
        """启动变量查询定时器"""
        print('_start_var_query_timer')
        if self.is_querying and self.status_handler and self.status_handler.is_connected():
            # 立即发送第一次查询
            self._send_query_var_message()

            # 设置定时器，每秒查询一次
            self.query_timer = threading.Timer(self.query_interval, self._query_var_periodically)
            self.query_timer.daemon = True
            self.query_timer.start()

    def _stop_var_query_timer(self):
        """停止变量查询定时器"""
        if self.query_timer:
            self.query_timer.cancel()
            self.query_timer = None

    def _query_var_periodically(self):
        """定时查询变量"""
        if self.is_querying and self.is_connected:
            try:
                # 发送查询消息
                self._send_query_var_message()
            except Exception as e:
                self.add_log(f"查询变量失败: {e}")

            # 重新设置定时器
            if self.is_querying:
                self.query_timer = threading.Timer(self.query_interval, self._query_var_periodically)
                self.query_timer.daemon = True
                self.query_timer.start()

    def select_model(self):
        """选择模型文件"""
        file_path = filedialog.askopenfilename(
            title="选择模型文件",
            filetypes=[("所有文件", "*.*")]
        )

        if file_path:
            filename = file_path.split("/")[-1]
            self.model_file_label.config(text=filename)
            self.add_log(f"选择了模型文件: {filename}")

    def download_model(self):
        """下载模型"""
        if not self.model_file_label.cget("text") or self.model_file_label.cget("text") == "未选择文件":
            messagebox.showwarning("警告", "请先选择模型文件")
            return

        self.download_status_label.config(text="下载中...")
        self.add_log("开始下载模型...")

        def simulate_download():
            time.sleep(2)
            self.root.after(0, lambda: self.download_status_label.config(text="已完成"))
            self.root.after(0, lambda: self.add_log("模型下载完成"))

        threading.Thread(target=simulate_download, daemon=True).start()

    def send_parameters(self, b_all_param=False):
        """
        下发参数

        Args:
            b_all_param: True表示发送所有参数，False表示只发送修改过的参数
        """
        if not self.is_connected or not self.ctrl_handler:
            messagebox.showwarning("警告", "请先连接到目标机")
            return

        # 检查是否有参数被修改
        has_modified = self._check_params_modified()

        if not has_modified:
            # 两种模式下，没有修改都提示并取消
            messagebox.showinfo("提示", "没有修改过的参数")
            self.add_log("没有修改过的参数，取消发送")
            return

        # 构建参数数据
        param_data = {}

        for i, param in enumerate(self.input_params):
            param_name = param.get("param", f"param{i + 1}")
            current_value = param.get("val", "")

            if b_all_param:
                # 发送所有参数
                param_data[param_name] = current_value
            else:
                # 只发送修改过的参数
                if i < len(self.original_input_params):
                    original_value = self.original_input_params[i].get("val", "")
                    if current_value != original_value:
                        param_data[param_name] = current_value
                else:
                    # 新增的参数
                    param_data[param_name] = current_value

        param_count = len(param_data)

        # 对于只发送修改的情况，再次检查（虽然前面检查过，但这里再做一次安全验证）
        if not b_all_param and param_count == 0:
            messagebox.showinfo("提示", "没有修改过的参数")
            self.add_log("没有修改过的参数，不发送")
            return

        # 添加日志说明发送模式
        if b_all_param:
            modified_count = self._get_modified_params_count() if hasattr(self, '_get_modified_params_count') else 0
            self.add_log(f"发送所有参数... 共{len(self.input_params)}个参数（其中{modified_count}个被修改）")
        else:
            self.add_log(f"发送修改过的参数... 共{param_count}个参数")

        # 构建JSON消息
        json_data = {
            "cmd": "SetParams",
            "count": param_count,
            "params": param_data
        }

        try:
            json_str = json.dumps(json_data, ensure_ascii=False)
            self.add_log(f"发送参数: {json_str}")

            # 通过控制链路发送参数
            if self.ctrl_handler.send_message(json_str):
                self.add_log(f"参数消息已发送 ({param_count}个参数)")

                # 记录参数数据
                try:
                    # 构建当前所有参数的数据字典
                    current_params = {}
                    for param in self.input_params:
                        param_name = param.get("param", "")
                        param_value = param.get("val", "")
                        current_params[param_name] = param_value

                    # 记录到Excel文件
                    self.record_parameters(current_params)
                except Exception as e:
                    self.add_log(f"记录参数数据失败: {e}")

                # 发送成功后，更新原始参数值
                self.original_input_params = [param.copy() for param in self.input_params]

                # 更新修改状态显示
                if hasattr(self, '_update_modified_status'):
                    self._update_modified_status()

            else:
                self.add_log("参数发送失败")

        except Exception as e:
            self.add_log(f"参数发送失败: {e}")

    def _update_original_params(self):
        """更新原始参数值为当前值"""
        self.original_input_params = [param.copy() for param in self.input_params]
        self.add_log("已更新参数修改记录")

    def reset_original_params(self):
        """重置原始参数为当前值"""
        self._update_original_params()
        self.add_log("已重置原始参数")

    def _check_params_modified(self):
        """检查是否有参数被修改（用于UI状态显示等）"""
        if len(self.input_params) != len(self.original_input_params):
            return True

        for i, param in enumerate(self.input_params):
            if i >= len(self.original_input_params):
                return True

            current_value = param.get("val", "")
            original_value = self.original_input_params[i].get("val", "")

            if current_value != original_value:
                return True

        return False

    def _get_modified_params_count(self):
        """获取被修改的参数数量"""
        count = 0

        for i, param in enumerate(self.input_params):
            if i >= len(self.original_input_params):
                count += 1
                continue

            current_value = param.get("val", "")
            original_value = self.original_input_params[i].get("val", "")

            if current_value != original_value:
                count += 1

        return count

    def _update_modified_status(self):
        """更新修改状态显示"""
        modified_count = self._get_modified_params_count()
        if modified_count > 0:
            self.modified_label.config(text=f"已修改: {modified_count}个", fg='red')
        else:
            self.modified_label.config(text="", fg='red')

    def reset_parameters(self):
        """复位所有参数到原始值"""
        if not self.input_params or not self.original_input_params:
            messagebox.showinfo("提示", "没有参数可复位")
            return

        # 统计复位数量
        reset_count = 0

        for i, param in enumerate(self.input_params):
            if i < len(self.original_input_params):
                current_value = param.get("val", "")
                original_value = self.original_input_params[i].get("val", "")

                # 如果值不同，进行复位
                if current_value != original_value:
                    param["val"] = original_value
                    reset_count += 1

                    # 更新表格显示
                    if i < len(self.param_rows):
                        self.param_rows[i]['value'].config(text=original_value)

        if reset_count > 0:
            # 更新表格显示
            self.update_params_table()

            # 记录日志
            self.add_log(f"已复位 {reset_count} 个参数到原始值")
            messagebox.showinfo("成功", f"已复位 {reset_count} 个参数")
        else:
            messagebox.showinfo("提示", "没有需要复位的参数")

        # 更新修改状态显示
        if hasattr(self, '_update_modified_status'):
            self._update_modified_status()

    def toggle_model_run(self):
        """切换模型运行状态"""
        if not self.model_file_label.cget("text") or self.model_file_label.cget("text") == "未选择文件":
            messagebox.showwarning("警告", "请先选择模型文件")
            return

        if not self.is_running:
            # 开始运行
            self.is_running = True
            self.is_querying = True
            self.run_button.config(text="模型停止", bg="lightcoral")
            self.start_time = time.time()
            self.stop_timer = False
            self.add_log("模型开始运行")
            # 记录初始参数
            try:
                # 构建初始参数数据字典
                init_params = {}
                for param in self.input_params:
                    param_name = param.get("param", "")
                    param_value = param.get("val", "")
                    init_params[param_name] = param_value

                # 记录到Excel文件
                self.record_parameters(init_params, "初始参数")
                self.add_log("已记录初始参数")
            except Exception as e:
                self.add_log(f"记录初始参数失败: {e}")

            # 启动变量查询定时器
            self._start_var_query_timer()

            # 启动计时器线程
            self.timer_thread = threading.Thread(target=self.update_timer, daemon=True)
            self.timer_thread.start()

            # 启动模拟数据更新线程
            # self.data_update_thread = threading.Thread(target=self.simulate_data_update, daemon=True)
            # self.data_update_thread.start()

        else:
            # 停止运行
            self.is_running = False
            self.is_querying = False
            self.run_button.config(text="模型运行", bg="SystemButtonFace")
            self.stop_timer = True
            self.time_label.config(text="00:00:00")

            # 停止查询定时器
            self._stop_var_query_timer()

            self.add_log("模型停止运行")

    def update_timer(self):
        """更新运行时间"""
        while self.is_running and not self.stop_timer:
            elapsed_time = int(time.time() - self.start_time)
            hours = elapsed_time // 3600
            minutes = (elapsed_time % 3600) // 60
            seconds = elapsed_time % 60
            time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

            self.root.after(0, lambda ts=time_str: self.time_label.config(text=ts))
            time.sleep(1)

    def simulate_data_update(self):
        """模拟数据更新 - 如果收到真实数据，可以注释掉这部分"""
        import random

        while self.is_running and not self.stop_timer:
            # 这里可以保留模拟数据更新，或者注释掉以便接收真实数据
            if not self.is_connected:  # 只有在没有连接时使用模拟数据
                if self.watch_variables:
                    # 随机更新一个变量的值
                    idx = random.randint(0, len(self.watch_variables) - 1)
                    var_type = self.watch_variables[idx].get("type", "int")

                    if var_type == "int":
                        new_val = str(random.randint(0, 100))
                    elif var_type == "float":
                        new_val = f"{random.uniform(0, 100):.2f}"
                    else:
                        new_val = f"val_{random.randint(100, 999)}"

                    self.watch_variables[idx]["val"] = new_val
                    self.root.after(0, self.update_watch_table)

            time.sleep(2)

    def add_log(self, message):
        """添加日志信息"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"

        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.log_text.update()

    def _send_heartbeat_message(self):
        if self.ctrl_handler and self.ctrl_handler.is_connected():
            # 构建心跳消息
            heartbeat_data = {
                "cmd": "Heart"
            }
            json_str = json.dumps(heartbeat_data)
            if self.ctrl_handler.send_message(json_str):
                self.add_log(f"发送心跳: {json_str}")
            else:
                self.add_log("心跳发送失败")
                # 发送失败，更新连接状态
                self._update_connection_status_display(False)

    def _update_connection_status_display(self, is_online, force_disconnect=False):
        """更新连接状态显示

        Args:
            is_online: 是否在线
            force_disconnect: 是否强制断开连接（避免递归调用）
        """
        if is_online:
            # 在线状态
            self.status_canvas.itemconfig(self.status_circle, fill='green')
            self.status_label.config(text="目标机在线", fg='green')
        else:
            # 离线状态
            self.status_canvas.itemconfig(self.status_circle, fill='red')
            self.status_label.config(text="目标机离线", fg='red')

            # 只有当不是强制断开时才进行断开操作
            if self.is_connected and not force_disconnect:
                # 设置 force_disconnect=True 避免递归调用
                self._disconnect_connections(force_status_update=True)

    def _start_heartbeat_mechanism(self):
        """启动心跳机制"""
        if not self.use_heartbeat:
            return

        # 停止现有的心跳定时器
        self._stop_heartbeat_mechanism()

        # 发送第一次心跳
        if self.is_connected and self.ctrl_handler and self.ctrl_handler.is_connected():
            self._send_heartbeat_message()

        # 启动心跳定时器
        if self.is_connected:
            self.heartbeat_timer = threading.Timer(self.heartbeat_interval, self._heartbeat_periodically)
            self.heartbeat_timer.daemon = True
            self.heartbeat_timer.start()

            # 启动状态检查定时器
            self._start_status_check_timer()

    def _stop_heartbeat_mechanism(self):
        """停止心跳机制"""
        if self.heartbeat_timer:
            self.heartbeat_timer.cancel()
            self.heartbeat_timer = None

        if self.status_check_timer:
            self.status_check_timer.cancel()
            self.status_check_timer = None

    def _heartbeat_periodically(self):
        """定时发送心跳"""
        if self.is_connected and self.use_heartbeat:
            try:
                if self.ctrl_handler and self.ctrl_handler.is_connected():
                    self._send_heartbeat_message()
            except Exception as e:
                self.add_log(f"发送心跳失败: {e}")

            # 重新设置定时器
            if self.is_connected and self.use_heartbeat:
                self.heartbeat_timer = threading.Timer(self.heartbeat_interval, self._heartbeat_periodically)
                self.heartbeat_timer.daemon = True
                self.heartbeat_timer.start()

    def _start_status_check_timer(self):
        """启动状态检查定时器"""
        if not self.use_heartbeat:
            return

        # 停止现有的状态检查定时器
        if self.status_check_timer:
            self.status_check_timer.cancel()

        # 启动新的状态检查定时器
        self.status_check_timer = threading.Timer(self.heartbeat_timeout, self._check_connection_status)
        self.status_check_timer.daemon = True
        self.status_check_timer.start()

    def _check_connection_status(self):
        """检查连接状态"""
        if not self.use_heartbeat or not self.is_connected:
            return

        current_time = time.time()

        # 检查是否超时
        if self.last_heartbeat_time is None or (current_time - self.last_heartbeat_time) > self.heartbeat_timeout:
            # 心跳超时，连接断开
            self.add_log("心跳超时，连接已断开")
            self.root.after(0, lambda: self._update_connection_status_display(False, force_disconnect=True))
            self.root.after(0, lambda: self.connect_button.config(text="连接", state="normal"))
        else:
            # 连接正常，继续检查
            self._start_status_check_timer()


if __name__ == "__main__":
    root = tk.Tk()
    app = HardwareSimulator(root)
    root.mainloop()
